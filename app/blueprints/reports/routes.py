"""
Blueprint: Laporan (EVENT-FIRST)
─────────────────────────────────
GET /api/reports         → REQ-REPORT-001  (Admin only)
GET /api/reports/export  → REQ-REPORT-002  (Admin only)

Perubahan dari versi sebelumnya:
  - tanggal sekarang OPSIONAL jika event_id disertakan.
  - Jika event_id ada tanpa tanggal → kembalikan SEMUA hari event tsb.
  - Jika event_id + tanggal → drill-down ke hari tertentu dalam event.
  - Jika hanya tanggal (tanpa event_id) → behavior lama (per-hari, default hari ini).
  - Response sekarang menyertakan tanggal_range: list tanggal unik yang punya data
    → dipakai frontend untuk menampilkan pill-selector per hari.
  - Export: tanggal opsional jika event_id ada; judul dokumen menyesuaikan scope.
"""

import io
import re as _re
from datetime import datetime, timedelta

from flask import Blueprint, jsonify, request, send_file

from app.extensions import supabase
from app.middleware.auth import admin_only

reports_bp = Blueprint("reports", __name__)


def _today_wib() -> str:
    """Tanggal hari ini dalam timezone WIB (UTC+7), format YYYY-MM-DD."""
    return (datetime.utcnow() + timedelta(hours=7)).strftime("%Y-%m-%d")


# ── GET /reports ──────────────────────────────────────────────────────────

@reports_bp.route("/reports", methods=["GET"])
@admin_only
def get_reports():
    tanggal  = (request.args.get("tanggal")  or "").strip()
    event_id = (request.args.get("event_id") or "").strip()

    # Fallback: jika tidak ada event_id maupun tanggal, default ke hari ini
    if not event_id and not tanggal:
        tanggal = _today_wib()

    try:
        query = (
            supabase.table("v_kunjungan_harian")
            .select("*")
            .order("waktu_masuk")
        )
        if event_id:
            query = query.eq("event_id", event_id)
        if tanggal:
            query = query.eq("tanggal", tanggal)

        result = query.execute()
        detail = result.data or []

        nama_event      = detail[0].get("nama_event") if detail else None
        actual_event_id = detail[0].get("event_id")   if detail else event_id or None

        # Kumpulkan tanggal unik yang punya data → untuk pill-selector di frontend
        tanggal_range = sorted({r["tanggal"] for r in detail if r.get("tanggal")})

        total_member = sum(1 for r in detail if r.get("tipe_pengunjung") == "member")
        total_biasa  = sum(1 for r in detail if r.get("tipe_pengunjung") == "biasa")

        return jsonify({
            "status": "success",
            "data": {
                "tanggal":       tanggal or None,
                "event_id":      actual_event_id,
                "nama_event":    nama_event,
                "tanggal_range": tanggal_range,
                "ringkasan": {
                    "total_kunjungan": len(detail),
                    "total_member":    total_member,
                    "total_biasa":     total_biasa,
                },
                "detail": detail,
            },
        }), 200

    except Exception:
        return jsonify({"status": "error", "message": "Terjadi kesalahan pada server"}), 500



# ── GET /reports/export ───────────────────────────────────────────────────

@reports_bp.route("/reports/export", methods=["GET"])
@admin_only
def export_report():
    fmt      = (request.args.get("format")   or "").lower()
    tanggal  = (request.args.get("tanggal")  or "").strip()
    event_id = (request.args.get("event_id") or "").strip()

    if fmt not in ("pdf", "excel"):
        return jsonify({
            "status": "error",
            "message": "Parameter format harus berupa 'pdf' atau 'excel'",
        }), 400

    # Minimal salah satu harus ada
    if not event_id and not tanggal:
        return jsonify({
            "status": "error",
            "message": "Parameter event_id atau tanggal harus disertakan",
        }), 400

    if tanggal:
        try:
            datetime.strptime(tanggal, "%Y-%m-%d")
        except ValueError:
            return jsonify({
                "status": "error",
                "message": "Format tanggal tidak valid. Gunakan YYYY-MM-DD",
            }), 400

    try:
        query = (
            supabase.table("v_kunjungan_harian")
            .select("*")
            .order("waktu_masuk")
        )
        if event_id:
            query = query.eq("event_id", event_id)
        if tanggal:
            query = query.eq("tanggal", tanggal)

        result     = query.execute()
        rows       = result.data or []
        nama_event = rows[0].get("nama_event", "-") if rows else "-"

        # Tentukan scope_label (subtitle dokumen) dan filename_suffix
        if tanggal:
            scope_label     = f"Tanggal: {tanggal}"
            filename_suffix = tanggal
        else:
            tanggal_unik = sorted({r["tanggal"] for r in rows if r.get("tanggal")})
            if not tanggal_unik:
                scope_label = "Seluruh Event"
            elif len(tanggal_unik) == 1:
                scope_label = f"Tanggal: {tanggal_unik[0]}"
            else:
                scope_label = f"Periode: {tanggal_unik[0]} s/d {tanggal_unik[-1]} ({len(tanggal_unik)} hari)"
            safe_name       = _re.sub(r"[^a-zA-Z0-9_-]", "_", nama_event)[:30]
            filename_suffix = safe_name

        if fmt == "excel":
            data, filename = _generate_excel(rows, scope_label, nama_event, filename_suffix)
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            data, filename = _generate_pdf(rows, scope_label, nama_event, filename_suffix)
            mime = "application/pdf"

        return send_file(
            io.BytesIO(data),
            mimetype=mime,
            as_attachment=True,
            download_name=filename,
        )
    except Exception:
        return jsonify({"status": "error", "message": "Terjadi kesalahan pada server"}), 500


# ── Helpers ───────────────────────────────────────────────────────────────

def _fmt_ts(ts) -> str:
    if not ts:
        return "-"
    return str(ts)[:19].replace("T", " ")


def _generate_excel(rows, scope_label, nama_event, filename_suffix):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Kunjungan"

    ws.merge_cells("A1:G1")
    ws["A1"] = "Laporan Kunjungan — Peken Banyumasan"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:G2")
    ws["A2"] = f"{nama_event}  |  {scope_label}"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.append([])

    headers = ["No", "Tipe", "Nama / Pengunjung", "Waktu Masuk", "Waktu Keluar", "Durasi (mnt)", "Status"]
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, row in enumerate(rows, start=1):
        even_fill = PatternFill("solid", fgColor="EFF6FF") if i % 2 == 0 else None
        values = [
            i,
            row.get("tipe_pengunjung", ""),
            row.get("nama_member") or "Pengunjung Biasa",
            _fmt_ts(row.get("waktu_masuk")),
            _fmt_ts(row.get("waktu_keluar")),
            row.get("durasi_menit") if row.get("durasi_menit") is not None else "-",
            row.get("status", ""),
        ]
        ws_row = ws.max_row + 1
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=ws_row, column=col, value=val)
            if even_fill:
                cell.fill = even_fill

    ws.append([])
    total_member = sum(1 for r in rows if r.get("tipe_pengunjung") == "member")
    total_biasa  = len(rows) - total_member
    ws.append([f"Total: {len(rows)}  |  Member: {total_member}  |  Pengunjung Biasa: {total_biasa}"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    for col, width in enumerate([5, 12, 28, 22, 22, 16, 12], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), f"laporan_kunjungan_{filename_suffix}.xlsx"


def _generate_pdf(rows, scope_label, nama_event, filename_suffix):
    from fpdf import FPDF

    pdf = FPDF(orientation="L", format="A4")
    pdf.set_margins(12, 15, 12)
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 9, "Laporan Kunjungan — Peken Banyumasan", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 7, f"{nama_event}  |  {scope_label}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    col_w   = [10, 20, 60, 48, 48, 30, 24]
    headers = ["No", "Tipe", "Nama / Pengunjung", "Waktu Masuk", "Waktu Keluar", "Durasi (mnt)", "Status"]
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(30, 58, 138)
    pdf.set_text_color(255, 255, 255)
    for w, h in zip(col_w, headers):
        pdf.cell(w, 8, h, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7.5)
    pdf.set_text_color(0, 0, 0)
    for i, row in enumerate(rows, start=1):
        fill = True
        if i % 2 == 0:
            pdf.set_fill_color(239, 246, 255)
        else:
            pdf.set_fill_color(255, 255, 255)
        nama   = (row.get("nama_member") or "Pengunjung Biasa")[:32]
        durasi = str(row.get("durasi_menit")) if row.get("durasi_menit") is not None else "-"
        pdf.cell(col_w[0], 7, str(i),                           border=1, align="C", fill=fill)
        pdf.cell(col_w[1], 7, row.get("tipe_pengunjung", ""),   border=1, align="C", fill=fill)
        pdf.cell(col_w[2], 7, nama,                             border=1, fill=fill)
        pdf.cell(col_w[3], 7, _fmt_ts(row.get("waktu_masuk")),  border=1, fill=fill)
        pdf.cell(col_w[4], 7, _fmt_ts(row.get("waktu_keluar")), border=1, fill=fill)
        pdf.cell(col_w[5], 7, durasi,                           border=1, align="C", fill=fill)
        pdf.cell(col_w[6], 7, row.get("status", ""),            border=1, align="C", fill=fill)
        pdf.ln()

    pdf.ln(4)
    total_member = sum(1 for r in rows if r.get("tipe_pengunjung") == "member")
    total_biasa  = len(rows) - total_member
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(0, 7, f"Total Kunjungan: {len(rows)}   |   Member: {total_member}   |   Pengunjung Biasa: {total_biasa}",
             new_x="LMARGIN", new_y="NEXT")

    return bytes(pdf.output()), f"laporan_kunjungan_{filename_suffix}.pdf"